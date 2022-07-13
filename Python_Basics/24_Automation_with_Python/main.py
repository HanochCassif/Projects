# In this project example script will open excel sheets
# then it will fix a column calculation and then it will copy this fixing to all rlevant rows

# it uses the transactions.xlsx as an input file

import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def workbook_process(filename):

    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    # max ros is 4. we want to read 3 rows (2 to 4 include) there for the below expression
    for row in range(2, sheet.max_row+1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value*0.9
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    # set a chart where top-left located on 'E2' cell
    chart = BarChart()
    chart.add_data(values)

    sheet..add_chart(chart, 'B7')
    wb.save(filename)


workbook_process('transactions.xlsx')